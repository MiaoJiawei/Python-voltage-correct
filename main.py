import numpy
import xlrd
import xlwt
import matplotlib.pyplot as plt


from scipy.optimize import leastsq
from openpyxl import load_workbook

# Request resistor value by least square.


def resistor(worksheet):
    # Define function.
    def hypothesis_func(w, x):
        w1, w0 = w
        return w1*x + w0

    # Build error function.
    def error_func(w, x, y):
        return hypothesis_func(w, x) - y

    # Pick maxium current value and voltage.
    def maxvalue(voltage, current):
        res = []
        cur_max = max(current)
        vol_max = voltage[current.index(max(current))]
        res.append(vol_max)
        res.append(cur_max)
        print(res)
        return res

    plt.figure(figsize=(8, 6))
    plt.grid(alpha=0.25)
    plt.xlabel('Voltage')
    plt.ylabel('Current')

    vol = []
    cur = []
    current_peaks = []
    voltage_peaks = []
    i = 0
    for col in worksheet.iter_cols(values_only=True):
        if i % 2 == 0:
            vol_temp = []
            j = 1
            for cell in col:
                if (j > 3) and (cell != None):
                    vol_temp.append(cell)
                j = j + 1
            vol = vol_temp
        else:
            cur_temp = []
            j = 1
            for cell in col:
                if (j > 3) and (cell != None):
                    cur_temp.append(cell)
                j = j + 1
            cur = cur_temp
            scrt = worksheet.cell(1, int( i+1 ))
            plt.plot(vol, cur, label='%s mV / Sec' %str(scrt.value))
            maxres = maxvalue(vol, cur)
            voltage_peaks.append(maxres[0])
            current_peaks.append(maxres[1])
        i = i + 1

    # Set initer value.
    print(voltage_peaks)
    print(current_peaks)
    voltage_raw = numpy.array(voltage_peaks)
    current_raw = numpy.array(current_peaks)
    w_init = [0.6, 1.5]

    # Fit function.
    fit = leastsq(error_func, w_init, args=(voltage_raw, current_raw))
    resistor = fit[0]
    print(resistor)
    plt.legend(loc='upper left')
    plt.show()
    return resistor[0]


def correct(voltage, current, resistor):
    voltage_correct = []
    for i in range(len(voltage)):
        voltage_correct.append(voltage[i] - (current[i] / resistor))

    return voltage_correct


wb = load_workbook('raw.xlsx')
ws_raw = wb.active
ws_cor = wb.create_sheet("Corrected")

resistance = resistor(ws_raw)

ws_cor.freeze_panes = 'A4'
i = 1
for raw in ws_raw.iter_rows(values_only=True):
    if i < 4:
        j = 1
        for cont in raw:
            ws_cor.cell(i, j).value = cont
            j = j + 1
        i = i + 1
    else:
        break

plt.figure(figsize=(8, 6))
plt.grid(alpha=0.25)
plt.xlabel('Voltage_correct')
plt.ylabel('Current')

vol = []
cur = []
voltage_corrected = []

i = 0
for col in ws_raw.iter_cols(values_only=True):
    if i % 2 == 0:
        vol_temp = []
        j = 1
        for cell in col:
            if (j > 3) and (cell != None):
                vol_temp.append(cell)
            j = j + 1
        vol = vol_temp
    else:
        cur_temp = []
        j = 1
        for cell in col:
            if (j > 3) and (cell != None):
                cur_temp.append(cell)
            j = j + 1
        cur = cur_temp
        voltage_corrected = correct(vol, cur, resistance)
        j = 4
        for cont in voltage_corrected:
            ws_cor.cell(j, i).value = cont
            j = j + 1
        j = 4
        for cont in cur:
            ws_cor.cell(j, (i+1)).value = cont
            j = j + 1  
        scrt = ws_raw.cell(1, int( i+1 ))
        plt.plot(voltage_corrected, cur, label='%s mV / Sec' %str(scrt.value))
    i = i + 1

# Plot the line.

# plt.plot(voltage, current, label='raw data')
# plt.plot(voltage_correct, current, label='corrected data')
plt.legend(loc='upper left')
plt.show()
# wb.save('raw.xlsx')