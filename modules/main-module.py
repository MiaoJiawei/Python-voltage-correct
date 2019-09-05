import matplotlib.pyplot as plt
import numpy
import scipy

from openpyxl import Workbook
from openpyxl import load_workbook

import functions

# Load the workbook.
wb = load_workbook('raw.xlsx')

# Set worksheets.
ws_raw = wb.active
ws_cor = wb.create_sheet("Corrected")
ws_fit_uc = wb.create_sheet('Fitted Non-diff')
ws_fit = wb.create_sheet('Fitted Non-diff Corrected')
ws_res = wb.create_sheet('Result')
ws_cut = wb.create_sheet("Cutted")
ws_cor_temp = wb.create_sheet("Corrected-Cutted")
ws_rearr = wb.create_sheet('Rearrangemented Corrected')
ws_rearr_temp = wb.create_sheet('Rearrangemented')

# Set header of table.
ws_cut = functions.set_cutted_header(ws_raw, ws_cut)
ws_cor = functions.set_common_header(ws_raw, ws_cor)
ws_cor_temp = functions.set_cutted_header(ws_raw, ws_cor_temp)
ws_rearr = functions.set_cutted_header(ws_raw, ws_rearr)
ws_rearr_temp = functions.set_cutted_header(ws_raw, ws_rearr_temp)
ws_fit = functions.set_cutted_header(ws_raw, ws_fit)
ws_fit_uc = functions.set_cutted_header(ws_raw, ws_fit_uc)

# Calculate resistance.
resistance_res = functions.resistor(ws_raw)
resistance = resistance_res[0][0]
peaks = resistance_res[1]

# Cut the CV curve.
cut = functions.cut(ws_raw, ws_cut)
ws_cut = cut[0]
voltage_range_cut = cut[1]

# Calculate the result vol-cur after corrected.
ws_cor_res = functions.correct(ws_raw, ws_cor, resistance)
ws_cor_temp_res = functions.correct(ws_cut, ws_cor_temp, resistance)
ws_cor = ws_cor_res[0]
ws_cor_temp = ws_cor_temp_res[0]

# Set voltage range.
voltage_range = [max(ws_cor_temp_res[2]), min(ws_cor_temp_res[1])]

# Rearrangements of data.
ws_rearr = functions.rearrangement(ws_cor_temp, ws_rearr, voltage_range)
ws_rearr_temp = functions.rearrangement(ws_cut, ws_rearr_temp, voltage_range_cut)

# Fit curves.
ws_fit = functions.fit_curve(ws_rearr, ws_fit)
ws_fit_uc =functions.fit_curve(ws_rearr_temp, ws_fit_uc)

# Write results.

# Remove template worksheets
# wb.remove(ws_cor_temp)

# Save file.
wb.save('res.xlsx')
