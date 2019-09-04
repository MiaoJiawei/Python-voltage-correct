import numpy

from scipy import interpolate
from openpyxl import load_workbook

import functions

wb = load_workbook('raw.xlsx')
ws_raw = wb.active
ws_cut = wb.create_sheet("Cutted")


def cut(vol, cur):
    vol_posi = []
    cur_posi = []
    vol_nega = []
    cur_nega = []

    for i in range(len(vol)):
        if vol[i-1] < vol[i]:
            vol_posi.append(vol[i])
            cur_posi.append(cur[i])
        else:
            vol_nega.append(vol[i])
            cur_nega.append(cur[i])

    return[vol_posi, cur_posi, vol_nega, cur_nega]

i = 0
for col in ws_raw.iter_cols(values_only=True):
    if i % 2 == 0:
        vol_temp = []
        j = 1
        for cell in col:
            if (j > 3) and (cell != None):
                vol_temp.append(cell)
            j = j + 1
        vol = vol_temp
    else:
        cur_temp = []
        j = 1
        for cell in col:
            if (j > 3) and (cell != None):
                cur_temp.append(cell)
            j = j + 1
        cur = cur_temp
        cut_res = cut(vol, cur)
        vol_posi = cut_res[0]
        cur_posi = cut_res[1]
        vol_nega = cut_res[2]
        cur_nega = cut_res[3]
        functions.pushpack(ws_cut, vol_posi, 2*i-1)
        functions.pushpack(ws_cut, cur_posi, 2*i)
        functions.pushpack(ws_cut, vol_nega, 2*i+1)
        functions.pushpack(ws_cut, cur_nega, 2*i+2)
    i = i + 1


wb.save('cut.xlsx')