from scipy import interpolate
from openpyxl import load_workbook

wb = load_workbook('raw.xlsx')
ws_raw = wb.active

i = 0
scrts = []
for col in ws_raw.iter_cols(values_only=True):
    if i % 2 == 0:
        pass
    else:
        scrt = ws_raw.cell(1, int( i+1 )).value
        scrts.append(scrt)
        locals()['a'+str(scrt)]=i
        print ('a'+str(scrt))
        print(eval('a'+str(scrt)))
    i = i + 1
print(scrts)
